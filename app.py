import random
import zipfile
from io import BytesIO

import streamlit as st
import pandas as pd
from datetime import datetime

# =====================================================================================
# 0. LOGIN GATE  (keeps the app private without ever putting a password on GitHub)
# =====================================================================================
# Two modes, tried in this order every time the app loads:
#   1) Streamlit secrets [auth] username/password (more secure, since the real
#      credentials never touch this .py file). Used automatically whenever a
#      .streamlit/secrets.toml (or Streamlit Cloud "Secrets") [auth] section is
#      configured.
#   2) A single hardcoded ADMIN_PASSWORD gate. Used automatically whenever [auth]
#      secrets are NOT configured — e.g. for quick local/admin use without setting
#      up secrets at all.
# .streamlit/secrets.toml, if used, should look like:
#   [auth]
#   username = "your_username"
#   password = "your_password"
ADMIN_PASSWORD = "kano"

# Playful Hinglish error messages for the hardcoded ADMIN_PASSWORD path, one is
# picked at random on every wrong attempt.
ADMIN_PASSWORD_ERROR_MESSAGES = [
    "Password इल्ले! 😅 इल्ले!, खम्मा घणी भाईसा, सॉरी। तुमसे सब कुछ हो पाएगा! यहां बहुत 🤪 दिमाग मत लगाओ, इस वेबसाइट को नहीं, 😂 इस गलत पासवर्ड को छोड़ दो!",
    "❌ Password इल्ले भाईसा! 😅 इल्ले! खम्मा घणी, सॉरी। तुम बाहुबली हो, तुमसे सब कुछ हो पाएगा! पर यहाँ फालतू 🤪 दिमाग मत लगाओ। अपनी सुंदर वेबसाइट को नहीं, 😂 इस सड़े हुए गलत पासवर्ड को छोड़ दो!",
    "❌ खम्मा घणी भाईसा, Password इल्ले! 😅 sorry! तुम तो मंगल ग्रह पर पानी खोज सकते हो, तुमसे सब कुछ हो पाएगा! पर यहाँ ज़्यादा 🤪 दिमाग मत लगाओ। इस सीधे-सादे वेबसाइट को नहीं, 😂 इस जाली पासवर्ड को छोड़ दो!",
    "❌ Password इल्ले! 😅 इल्ले! खम्मा घणी भाईसा, सॉरी। लोड मत लो, तुमसे सब कुछ हो पाएगा! पर यहाँ फालतू 🤪 दिमाग मत लगाओ। दुनिया छोड़ दो, मोक्ष पकड़ लो, पर पहले 😂 इस गलत पासवर्ड को छोड़ दो!",
    "❌ अरे भाईसा! Password इल्ले! 😅 खम्मा घणी, सॉरी। तुम चाहो तो सिस्टम हिला सकते हो, तुमसे सब कुछ हो पाएगा! पर यहाँ ज़्यादा 🤪 दिमाग मत लगाओ। इस निर्दोष वेबसाइट को नहीं, 😂 इस भूतिया गलत पासवर्ड को छोड़ दो!",
]


def check_login():
    if st.session_state.get("authenticated", False):
        return True

    auth_cfg = st.secrets.get("auth", None)

    # ---- Mode 1: secrets-based username/password ----
    if auth_cfg:
        st.title("🔒 Breakout List Merger — Login")
        with st.form("login_form"):
            user = st.text_input("Username")
            pwd = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Log in")
        if submitted:
            if user == auth_cfg.get("username") and pwd == auth_cfg.get("password"):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Invalid username or password.")
        return False

    # ---- Mode 2: hardcoded ADMIN_PASSWORD gate ----
    st.markdown(
        "<p style='text-align: center; margin-top: 100px; color: Green; font-size: 18px;'>"
        "📊 Breakout List — Merge All Sheets Into One</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<h1 style='text-align: center; margin-top: 0px; font-size: 20px;'>🔐 Admin Login</h1>",
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        with st.form("admin_login_form"):
            pwd = st.text_input("Enter Password", type="password")
            submit = st.form_submit_button("Login", use_container_width=True)
            if submit:
                if pwd == ADMIN_PASSWORD:
                    st.session_state["authenticated"] = True
                    st.rerun()
                else:
                    st.error(random.choice(ADMIN_PASSWORD_ERROR_MESSAGES))

    dynamic_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    st.markdown(
        f"<p style='text-align: center; color: gray; font-size: 14px; margin-top: 20px;'>"
        f"Data refreshed: {dynamic_time}</p>",
        unsafe_allow_html=True,
    )
    return False


st.set_page_config(page_title="Breakout List Merger", page_icon="📊", layout="wide")

# ==========================================
# 🛡️ HIDE GITHUB ICON ONLY
# ==========================================
hide_github_icon = """
<style>
    [data-testid="stToolbar"] {
        right: 2rem;
    }
    [data-testid="stToolbar"]::before {
        content: "";
    }
    button[kind="header"] {display: none;}
</style>
"""
st.markdown(hide_github_icon, unsafe_allow_html=True)

if not check_login():
    st.stop()

st.title("📊 Breakout List — Merge All Sheets Into One")
st.write(
    "Upload all your Breakout List Excel files (Sheet-1 to Sheet-7, or any number of files). "
    "The app will combine every sheet from every file into a single merged 'Top 250 Stocks' tab, "
    "remove duplicate rows, and let you download the final combined file. "
    "You can also upload .zip files (e.g. a zipped 'Cumulative Average2' workbook) — "
    "any recognized pass-through sheet inside is added as its own extra tab, completely unchanged."
)

uploaded_files = st.file_uploader(
    "Upload Excel files (.xlsx) or zipped Excel files (.zip)",
    type=["xlsx", "zip"],
    accept_multiple_files=True,
    help="You can select and drop multiple files at once, including .zip files containing .xlsx files.",
)

# Optional: let user decide whether to drop exact duplicate rows
drop_duplicates = st.checkbox("Remove duplicate rows (exact matches)", value=True)


# Sheet names that should be added to the final workbook as their OWN extra tab,
# copied exactly as-is (values/formulas unchanged) — NOT merged into "Top 250 Stocks".
# Add more names here if other pass-through workbooks come up in future.
PASSTHROUGH_SHEET_NAMES = {"Cumulative Average2"}


class NamedBytesIO(BytesIO):
    """A BytesIO that also carries a .name, so it behaves like an uploaded file
    for pandas/openpyxl (which expect an object with a .name attribute)."""

    def __init__(self, data: bytes, name: str):
        super().__init__(data)
        self.name = name


def expand_uploads(files):
    """Take the raw list of uploaded files (.xlsx and/or .zip) and return a flat
    list of NamedBytesIO objects, one per .xlsx file — extracting any .xlsx found
    inside uploaded .zip archives."""
    expanded = []
    for f in files:
        lower_name = f.name.lower()
        if lower_name.endswith(".zip"):
            try:
                with zipfile.ZipFile(f) as zf:
                    for entry in zf.namelist():
                        if entry.lower().endswith(".xlsx") and not entry.startswith("__MACOSX"):
                            data = zf.read(entry)
                            display_name = f"{f.name} → {entry.split('/')[-1]}"
                            expanded.append(NamedBytesIO(data, display_name))
            except Exception as e:
                st.error(f"Could not read zip file {f.name}: {e}")
        else:
            expanded.append(f)
    return expanded


def split_breakout_and_passthrough(files):
    """Classify each uploaded (expanded) .xlsx file:
    - If ANY of its sheets is a known pass-through sheet (see PASSTHROUGH_SHEET_NAMES),
      pull that sheet out untouched to be added as its own tab later.
    - Any remaining sheets in that same file, plus every sheet from files with no
      pass-through match, are treated as normal breakout-list input for merging.
    Returns: (breakout_files, passthrough_sheets)
      breakout_files: list of NamedBytesIO, ready for load_and_combine()
      passthrough_sheets: list of (sheet_name, openpyxl Worksheet) to copy as-is
    """
    import openpyxl

    breakout_files = []
    passthrough_sheets = []

    for f in files:
        try:
            f.seek(0)
            raw = f.read()
            f.seek(0)
            wb = openpyxl.load_workbook(BytesIO(raw), data_only=False)
            matched_any = False
            for sheet_name in wb.sheetnames:
                if sheet_name in PASSTHROUGH_SHEET_NAMES:
                    passthrough_sheets.append((sheet_name, wb[sheet_name]))
                    matched_any = True
            if not matched_any:
                # No pass-through sheet found — treat whole file normally
                f.seek(0)
                breakout_files.append(f)
        except Exception as e:
            st.error(f"Could not inspect {f.name}: {e}")

    return breakout_files, passthrough_sheets


def copy_passthrough_sheet(target_wb, sheet_name, source_ws):
    """Copy a worksheet's cells (values/formulas + basic formatting) exactly
    as-is into target_wb, under the same sheet name, with no data changes."""
    target_ws = target_wb.create_sheet(title=sheet_name)

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.alignment = cell.alignment.copy()

    # Preserve column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        if dim.width:
            target_ws.column_dimensions[col_letter].width = dim.width

    # Preserve merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))


def load_and_combine(files):
    """Read every sheet of every uploaded file and stack them into one DataFrame."""
    all_frames = []
    file_summary = []

    for f in files:
        try:
            xls = pd.ExcelFile(f)
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if df.empty:
                    continue
                df["Source File"] = f.name
                df["Source Sheet"] = sheet_name
                all_frames.append(df)
                file_summary.append(
                    {"File": f.name, "Sheet": sheet_name, "Rows": len(df)}
                )
        except Exception as e:
            st.error(f"Could not read {f.name}: {e}")

    if not all_frames:
        return None, pd.DataFrame()

    combined = pd.concat(all_frames, ignore_index=True, sort=False)

    # Drop columns that shouldn't appear in the final combined output
    columns_to_drop = ["Source File", "Source Sheet", "Date"]
    combined = combined.drop(columns=[c for c in columns_to_drop if c in combined.columns])

    # Rename columns for the final output
    combined = combined.rename(columns={"Stock": "Symbol"})

    return combined, pd.DataFrame(file_summary)


# Each entry: (base_url, suffix, display_prefix, uses_symbol_in_formula)
# uses_symbol_in_formula = True  -> build a live =HYPERLINK(...) formula that
#                                    references that row's own Symbol cell
#                                    (behaves like a normal dragged-down formula)
# uses_symbol_in_formula = False -> static display only (no live formula)
HYPERLINK_SPECS = {
    "NSE Chart": (
        "https://www.nseindia.com/get-quotes/equity?symbol=",
        "",
        "n ",
        True,
    ),  # Static display
    "Trading View": (
        "https://www.tradingview.com/symbols/",
        "",
        "Tre ",
        True,
    ),
    "History Data": (
        "https://www.equitypandit.com/historical-data/",
        "",
        "his ",
        True,
    ),  # Replace with actual base URL
    "Screener": (
        "https://www.screener.in/company/",
        "",
        "Scr ",
        True,
    ),  # Replace with actual base URL structure
    "Zerodha": (
        "https://zerodha.com/markets/stocks/NSE/",
        "",
        "Z ",
        True,
    ),  # Replace with actual base URL
    "Chartlink": (
        "https://chartink.com/stocks/",
        ".html",
        "CL ",
        True,
    ),  # Replace with actual base URL
    "Marketsmith": (
        "https://marketsmithindia.com/mstool/eval/",
        "/evaluation.jsp",
        "ms ",
        True,
    ),  # Replace with actual base URL
}


def to_excel_bytes(df: pd.DataFrame, passthrough_sheets=None) -> bytes:
    """Convert a DataFrame to an in-memory Excel file (a 'Top 250 Stocks' tab),
    with live HYPERLINK formulas added per row for each configured site.
    Any (sheet_name, worksheet) pairs in passthrough_sheets are appended as
    additional tabs, copied exactly as-is with no changes."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if df is not None and len(df) > 0:
            df.to_excel(writer, index=False, sheet_name="Top 250 Stocks")

            worksheet = writer.sheets["Top 250 Stocks"]
            headers = list(df.columns)

            # Auto-fit original column widths for readability
            for i, col in enumerate(headers, start=1):
                max_len = max(
                    df[col].astype(str).map(len).max() if len(df) else 0, len(str(col))
                )
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=i).column_letter
                ].width = min(max_len + 2, 40)

            # Add hyperlink columns (NSE Chart, Trading View, History Data, Screener,
            # Zerodha, Chartlink, Marketsmith), each written as a real Excel formula
            # per row, referencing that row's own Symbol cell.
            if "Symbol" in headers and len(df) > 0:
                symbol_col_idx = headers.index("Symbol") + 1  # 1-based
                symbol_col_letter = worksheet.cell(row=1, column=symbol_col_idx).column_letter

                start_col = len(headers) + 1
                for offset, (label, (base_url, suffix, prefix, use_formula)) in enumerate(
                    HYPERLINK_SPECS.items()
                ):
                    col_idx = start_col + offset
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter

                    # Header
                    header_cell = worksheet.cell(row=1, column=col_idx, value=label)
                    header_cell.font = header_cell.font.copy(bold=True)

                    for row_num in range(2, len(df) + 2):
                        sym_ref = f"{symbol_col_letter}{row_num}"
                        cell = worksheet.cell(row=row_num, column=col_idx)
                        if use_formula:
                            # e.g. =HYPERLINK("https://.../"&A2&".html","CL "&A2)
                            formula = (
                                f'=HYPERLINK("{base_url}"&{sym_ref}&"{suffix}",'
                                f'"{prefix}"&{sym_ref})'
                            )
                            cell.value = formula
                        else:
                            # Static display only (no live formula), e.g. NSE Chart
                            cell.value = prefix

                    worksheet.column_dimensions[col_letter].width = max(len(label) + 2, 14)

        # Append pass-through sheets (e.g. "Cumulative Average2"), unchanged.
        if passthrough_sheets:
            target_wb = writer.book
            for sheet_name, source_ws in passthrough_sheets:
                # Avoid name clashes if the same sheet name shows up twice
                final_name = sheet_name
                suffix_n = 2
                while final_name in target_wb.sheetnames:
                    final_name = f"{sheet_name} ({suffix_n})"
                    suffix_n += 1
                copy_passthrough_sheet(target_wb, final_name, source_ws)

    return output.getvalue()



if uploaded_files:
    expanded_files = expand_uploads(uploaded_files)
    breakout_files, passthrough_sheets = split_breakout_and_passthrough(expanded_files)

    combined_df, summary_df = (None, pd.DataFrame())
    if breakout_files:
        combined_df, summary_df = load_and_combine(breakout_files)

    if combined_df is not None:
        original_count = len(combined_df)

        if drop_duplicates:
            # Drop duplicates based on actual data columns only (ignore source tracking columns)
            data_cols = [
                c for c in combined_df.columns if c not in ("Source File", "Source Sheet")
            ]
            combined_df = combined_df.drop_duplicates(subset=data_cols, keep="first")

        st.success(
            f"Merged {len(breakout_files)} breakout file(s) → {len(summary_df)} sheet(s) → "
            f"{original_count} rows read, {len(combined_df)} rows in final 'Top 250 Stocks' tab."
        )

        with st.expander("📄 File / Sheet summary"):
            st.dataframe(summary_df, use_container_width=True)

        st.subheader("Top 250 Stocks Preview")
        st.dataframe(combined_df, use_container_width=True)

    if passthrough_sheets:
        st.success(
            f"Added {len(passthrough_sheets)} pass-through tab(s) unchanged: "
            + ", ".join(name for name, _ in passthrough_sheets)
        )

    if combined_df is not None or passthrough_sheets:
        excel_bytes = to_excel_bytes(combined_df, passthrough_sheets)
        today_str = datetime.now().strftime("%Y-%m-%d")

        st.download_button(
            label="⬇️ Download Combined Excel File",
            data=excel_bytes,
            file_name=f"Breakout_List_Combined_{today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.warning("No readable data found in the uploaded file(s).")
else:
    st.info("👆 Upload one or more Excel files to get started.")
